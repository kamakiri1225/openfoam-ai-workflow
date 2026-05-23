# baseからFieldファイルをコピーする
import os
import shutil
from OpenFOAMlibs.boundary_condition import boundary_def
import openpyxl
 

def setting_BoundaryCondition(boundary_file):
    wb = openpyxl.load_workbook(boundary_file)
    
    ws = wb['境界条件指定']
    boundary_list = []

    for i in range(2,ws.max_row + 1):
        boundary_dict = {}
        value_list = []
        if ws.cell(row=i, column=2).value == None:
            break
        for j in range(4,8):
            value_list.append(ws.cell(row=i, column=j).value)
        boundary_dict['patchName'] = ws.cell(row=i, column=2).value
        boundary_dict['bounaryType'] = ws.cell(row=i, column=3).value
        boundary_dict['value'] = value_list
        boundary_list.append(boundary_dict)

    print(boundary_list)

    # 境界条件の設定
    PWD = os.getcwd() # 現在のディレクトリパス
    for boundary_ in boundary_list:
        if boundary_['patchName'].split('_')[1] != "wall":
            print(boundary_)
            boundary_def(PWD,boundary_)
