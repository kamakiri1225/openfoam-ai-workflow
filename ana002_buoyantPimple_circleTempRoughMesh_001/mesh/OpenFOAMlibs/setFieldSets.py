import glob
import os
from PyFoam.RunDictionary.ParsedParameterFile import ParsedParameterFile
from os import path
import openpyxl
from OpenFOAMlibs.subSetFields import *

def isSetField(boundary_file):
        # エクセルから境界条件の読み込み
    wb = openpyxl.load_workbook(boundary_file)
    ws = wb['境界条件指定']
    flag = ws.cell(row=1, column=11).value

    if flag == "yes":
        return True
    else:
        return False

def inputRegionDict(boundary_file):
    wb = openpyxl.load_workbook(boundary_file)
    ws = wb['境界条件指定']

    boxToCell_List = [ws.cell(row=3, column=11 + i).value for i in range(6)]
    boxToCell_Dict = {
                    'regions' : boxToCell_List,
                    'var' : ws.cell(row=3, column=18).value
    }

    cylinderToCell_List = [ws.cell(row=5, column=11 + i).value for i in range(6)]
    cylinderToCell_Dict = {
                    'regions' : cylinderToCell_List,
                    'radius' : ws.cell(row=5, column=17).value,
                    'var' : ws.cell(row=5, column=18).value
    }
    setFields_dic = {
        'boxToCell': boxToCell_Dict,
        'cylinderToCell' :cylinderToCell_Dict
    }
    print(setFields_dic)
    return setFields_dic

def createRegion(setFieldsParsedParameterFile, input_dict):
    setFields_set_fuc(setFieldsParsedParameterFile, input_dict) #setFieldの設定